# IMPORTS:

from openpyxl import Workbook, load_workbook
from csv import reader
from os.path import exists
from glob import glob
from os import chdir, getcwd
from sys import platform
from Attributes import path_to_csv_dir, teacher, min_percent, date_line_sub_str, last_line_sub_str, path_to_register, names_heading, include_year, attrs, most_wanted, get_students, get_date
from tkinter import Tk, LabelFrame, Button, Label, StringVar, Entry, OptionMenu
from tkinter.messagebox import askyesno, showerror
from tkinter.filedialog import askdirectory


# FUNCTIONS:

def open_file_or_directory(path):
    if platform == "win32":
        from os import startfile
        startfile(path)  # Windows
    elif platform == "darwin":  # macOS
        from subprocess import run
        run(["open", path])
    else:  # Linux and others
        from subprocess import run
        run(["xdg-open", path])


def input_dir():
    """Inputs the CSVs' directory."""
    path = askdirectory(initialdir='.', mustexist=True, title='Please select the CSVs\' directory.')
    if path:
        global path_to_csv_dir
        path_to_csv_dir = path
        dir_button.configure(text=path_to_csv_dir.rsplit('/', maxsplit=1)[-1])


def main_process():
    """Main Process."""

    global path_to_register, teacher, min_percent

    teacher = teacher_var.get().strip()
    if not teacher:
        showerror(title='Teacher\'s Name Unfilled', message='Please enter the teacher\'s name.')
        return
    min_percent = int(min_percent_var.get())
    print(f'\n{path_to_csv_dir}\n{teacher}\n{min_percent}')  # debugging

    # CONNECTING TO ATTENDANCE REGISTER:

    if exists(path_to_register) and askyesno(title=f'"{path_to_register}" Already Exists', message='Continue appending attendance in it only? \n(Yes: Continue in the same register \nNo: Continue in a new register)'):

        from Attributes import heading_row, start_column

        wb = load_workbook(path_to_register)
        sheet = wb.active
        # print(sheet)  #debugging

        # Calculating column number of the names heading in the attendance register:
        reg_names_col = 1
        while sheet.cell(row=heading_row, column=reg_names_col).value.strip() != names_heading:
            reg_names_col += 1

        if not start_column:
            start_column = sheet.max_column + 1

    else:  # creating

        i = 1
        copy = path_to_register
        while exists(path_to_register):
            path_to_register = f'{copy[:-5]} ({i}).xlsx'
            i += 1
        del i, copy
        print('\nPath to register:', path_to_register)

        wb = Workbook()
        wb.save(path_to_register)
        sheet = wb.active
        # print(sheet)  #debugging

        # Inserting some basic required values in attendance register:
        sheet.cell(row=1, column=1, value='TITLE')

        heading_row, start_column = 3, 3

        sheet.cell(row=heading_row, column=1, value='R.N.')
        sheet.cell(row=heading_row, column=2, value=names_heading)
        reg_names_col = 2

        # Inserting student names:
        for rn, name in enumerate(get_students(), start=heading_row+1):
            sheet.cell(row=rn, column=1, value=rn-heading_row)
            sheet.cell(row=rn, column=2, value=name)

        # wb.save(path_to_register); exit()  #debugging

    # MAIN:

    def get_reg_name(rn_):
        """Returns the name in attendance register for roll number "rn_"."""
        cell_value_ = sheet.cell(row=rn_, column=reg_names_col).value
        return cell_value_ and cell_value_.strip()

    cwd = getcwd()  # storing cwd
    chdir(path_to_csv_dir)

    csv_list = glob('*.csv')
    if len(csv_list) == 0:
        showerror(title='No CSVs Found', message=f'No CSVs found in "{path_to_csv_dir}", add the CSVs and try again.')
        return

    day = start_column

    for i, csv_file in enumerate(sorted(csv_list), start=1):
        print(f'\n{i})', csv_file)

        with open(csv_file, newline='', encoding='utf-8-sig') as file:
            # encoding='utf-8-sig' -> https://stackoverflow.com/questions/34399172/why-does-my-python-code-print-the-extra-characters-%c3%af-when-reading-from-a-tex

            data = reader(file)

            date = get_date(filename=csv_file)
            # Verifying Date:
            while True:
                try:
                    cell = next(data)[0]  # taking only one cell from every row
                except IndexError:  # (empty row)
                    continue
                except StopIteration:  # reached the end of the csv
                    showerror(title='An Error Occurred', message=f'''"date_line_sub_str" didn't match... Please correct it in "{attrs}" and run the program again. \n\nEXITING WITHOUT SAVING''')
                    exit()
                # print(cell)  #debugging

                if date_line_sub_str in cell:
                    if str(date) not in cell:
                        showerror(title='Something Went Wrong', message=f'''Date in CSV filename ("{csv_file}") and inside CSV ("{cell}") didn't match... Please make sure it's same (else wrong attendance will be marked) and run the program again. \n\nEXITING WITHOUT SAVING''')
                        exit()
                    break

            # Skipping to names-starting:
            while True:
                try:
                    if last_line_sub_str in next(data)[0]:
                        break
                except IndexError:  # (empty row)
                    continue
                except StopIteration:  # reached the end of the csv
                    showerror(title='An Error Occurred', message=f'''"last_line_sub_str" didn't match... Please correct it in "{attrs}" and run the program again. \n\nEXITING WITHOUT SAVING''')
                    exit()

            # MARKING ATTENDANCE:

            sheet.cell(row=heading_row, column=day, value=date.strftime(format='%d-%m-%Y' if include_year else '%d-%m'))  # date heading

            roll_no = heading_row + 1

            # print(list(data)); exit()  #debugging
            attendees = list(filter(lambda x: x, sorted(map(lambda x: x[0].strip().title() if x != [] else '', data))))  # doesn't matter what's outside, the data will be formatted well after coming inside the program! (title cased etc.)
            # print(attendees)  #debugging
            if len(attendees) != len(set(attendees)):  # duplicate names check
                from collections import Counter
                histogram = Counter(attendees)
                raise ValueError(f'"{csv_file}" contains duplicate names {list(filter(lambda x: histogram[x] > 1, histogram))}, please fix it and run the program again.')

            for attendee in attendees:

                if attendee == teacher:  # do nothing
                    continue

                while True:
                    reg_name = get_reg_name(roll_no)

                    if reg_name == attendee:
                        sheet.cell(row=roll_no, column=day, value='P')
                        roll_no += 1  # next roll no.
                        break

                    if (not reg_name) or (reg_name > attendee):
                        print("Name didn't match:", attendee)
                        break

                    sheet.cell(row=roll_no, column=day, value='A')
                    roll_no += 1  # next roll no.

            while get_reg_name(roll_no):  # marking absent of remaining students
                sheet.cell(row=roll_no, column=day, value='A')
                roll_no += 1  # next roll no.

        day += 1  # attendance of next day

    # ATTENDANCE TOTAL:

    sheet.cell(row=heading_row, column=day, value='Total')

    low_attendees = best_attendees = []
    best = best_percent = 0

    for rn in range(heading_row+1, roll_no):

        totals = 0  # no. of columns of 'Total'
        count = 0
        for j in range(reg_names_col+1, day):
            cell_value = sheet.cell(row=rn, column=j).value  # 'P' / 'A' / int

            if isinstance(cell_value, int) or isinstance(cell_value.strip(), int):
                totals += 1
                continue

            if cell_value.strip() == 'P':
                count += 1

        sheet.cell(row=rn, column=day, value=count)

        attendance_percent = count / (day-(reg_names_col+1)-totals) * 100
        if attendance_percent < min_percent:
            low_attendees.append({'RN': rn-heading_row, 'Name': get_reg_name(rn), 'Attendance %': float('{:.2f}'.format(attendance_percent))})

        if count > best:
            best = count
            best_percent = attendance_percent
            best_attendees = [rn]
        elif count == best:
            best_attendees.append(rn)

    chdir(cwd)  # changing back so that files save in the right place

    with open(file=most_wanted, mode='a') as file:

        file.write(f'For "{path_to_register}" till date "{date}":\n')

        if len(low_attendees) > 0:
            file.write(f'\nStudent{"" if len(low_attendees)==1 else "s"} with low attendance (< {min_percent}%): \n')
            file.writelines(map(lambda x: str(x)+'\n', low_attendees))

        if len(best_attendees) > 0:
            file.write(f'\nStudent{"" if len(best_attendees)==1 else "s"} with best attendance ({float("{:.2f}".format(best_percent))}%): \n')
            file.writelines(map(lambda rn_: str({'RN': rn_-heading_row, 'Name': get_reg_name(rn_)})+'\n', best_attendees))

        file.write('\n' + '#'*78 + '\n\n')

    try:
        wb.save(path_to_register)
    except PermissionError as e:
        exit(f'Error Occurred: {e} \nNOTHING HAS BEEN SAVED \nMake sure the \'{path_to_register}\' is not opened in some other program.')

    if askyesno(title='SUCCESS!', message=f'Open the files? ("{path_to_register}" & "{most_wanted}")'):
        open_file_or_directory(path_to_register)
        open_file_or_directory(most_wanted)

    window.destroy()


# GUI MAIN:

window = Tk()
window.title('Attendance Automator')

frame = LabelFrame(window, text='Please enter:', padx=32, pady=32)
frame.grid()

Label(master=frame, text='CSVs\' Directory: ').grid(row=0, column=0)

dir_button = Button(master=frame, text=path_to_csv_dir, command=input_dir)
dir_button.grid(row=0, column=1)

Label(master=frame, text='Teacher\'s Name: ').grid(row=1, column=0)

teacher_var = StringVar()
teacher_var.set(teacher)
teacher_button = Entry(master=frame, textvariable=teacher_var)
teacher_button.grid(row=1, column=1)

Label(master=frame, text='Minimum Percent: ').grid(row=2, column=0)

min_percent_var = StringVar()
min_percent_var.set(str(min_percent))
min_percent_om = OptionMenu(frame, min_percent_var, *range(50, 75+1))
# min_percent_om.configure(direction='flush')
min_percent_om.grid(row=2, column=1)

start_button = Button(master=frame, text='Start Process', command=main_process)
start_button.grid(row=3, columnspan=2)


# MAINLOOP:

window.mainloop()
